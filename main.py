"""
Employee Agent POC
- AI Phone Call : Twilio (outbound call → records response as MP3)
                  OpenAI Whisper (downloads MP3 → accurate voice-to-text)
- AI Email      : Gmail SMTP via smtplib (built-in, no extra library)
- Output        : Writes new columns back into input Excel:
                    Status               → 'DisQualified' (born >2000, not terminated)
                    Email_Sent           → Yes / No
                    Email_Text           → body of email sent
                    Phone_Call_Status    → Completed / No Answer / Not Applicable
                    Phone_Conversation   → voice-to-text transcript via Whisper

Setup:
    pip install twilio groq pandas openpyxl requests

Usage:
    python poc_agent.py               # run full flow
    python poc_agent.py --test-email  # test just email
    python poc_agent.py --test-call   # test just call
"""

import os
import sys
import time
import smtplib
import requests
import pandas as pd
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────
# CONFIG  ← set as env vars or hardcode here for POC
# ─────────────────────────────────────────────────────────────────
TWILIO_ACCOUNT_SID  = os.getenv("TWILIO_ACCOUNT_SID",  "AC79aa99f1806978009e27aaa059228aa9")
TWILIO_AUTH_TOKEN   = os.getenv("TWILIO_AUTH_TOKEN",   "fd943dd5fdbc626701d790c6c306dc36")
TWILIO_FROM_NUMBER  = os.getenv("TWILIO_FROM_NUMBER",  "+19152283121")   # your Twilio number

GMAIL_ADDRESS       = os.getenv("GMAIL_ADDRESS",       "lakshmideepakb@gmail.com")
GMAIL_APP_PASSWORD  = os.getenv("GMAIL_APP_PASSWORD",  "jxtk wadn dvoy guvc")  # 16-char app password

GROQ_API_KEY        = os.getenv("GROQ_API_KEY")

# ── Test targets ──────────────────────────────────────────────────
TEST_PHONE = "+919176960154"
TEST_EMAIL = "lakshmideepakb@gmail.com"

INPUT_FILE  = "sample_data.xlsx"
OUTPUT_FILE = "sample_data.xlsx"   # write results back into the same input file

BIRTH_CUTOFF     = datetime(2000, 1, 1)
TERM_DATE_CUTOFF = datetime(2023, 1, 1)


# ─────────────────────────────────────────────────────────────────
# GROQ WHISPER TRANSCRIPTION  (free, fast, accurate for Indian English)
# ─────────────────────────────────────────────────────────────────
def transcribe_with_whisper(recording_url: str) -> str:
    """
    Downloads the Twilio MP3 recording and transcribes using Groq Whisper.
    Groq is FREE, fast (~instant), and uses whisper-large-v3 (most accurate model).

    Steps:
      1. Download MP3 from Twilio (requires Twilio auth)
      2. POST audio file to Groq transcription API
      3. Return the transcript text
    """
    from groq import Groq
    client = Groq(api_key=GROQ_API_KEY)

    print("     Downloading recording from Twilio...")
    response = requests.get(
        recording_url,
        auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
        timeout=30
    )

    if response.status_code != 200:
        print(f"     ❌ Failed to download recording: HTTP {response.status_code}")
        return "Recording download failed"

    tmp_file = "tmp_recording.mp3"
    with open(tmp_file, "wb") as f:
        f.write(response.content)
    print(f"     Recording downloaded ({len(response.content) // 1024} KB)")

    print("     Sending to Groq Whisper for transcription...")

    # Retry up to 3 times — handles Groq 503 temporary overload spikes
    for attempt in range(1, 4):
        try:
            with open(tmp_file, "rb") as audio_file:
                result = client.audio.transcriptions.create(
                    model="whisper-large-v3",   # best accuracy, completely free on Groq
                    file=audio_file,
                    language="en"
                )
            transcript = result.text
            print(f"     ✅ Groq Whisper transcript: {transcript}")
            return transcript

        except Exception as e:
            if "503" in str(e) and attempt < 3:
                wait = attempt * 10   # wait 10s, then 20s before retrying
                print(f"     ⚠️  Groq unavailable (attempt {attempt}/3), retrying in {wait}s...")
                time.sleep(wait)
            else:
                print(f"     ❌ Groq Whisper error: {e}")
                return f"Transcription error: {e}"

        finally:
            if os.path.exists(tmp_file):
                os.remove(tmp_file)


# ─────────────────────────────────────────────────────────────────
# CALL AGENT
# ─────────────────────────────────────────────────────────────────
def make_call_and_transcribe(to_number: str, person_name: str) -> dict:
    """
    1. Places outbound Twilio call
    2. Plays HR prompt → records spoken response (finishOnKey="" prevents keypress ending it)
    3. Downloads the MP3 recording
    4. Transcribes via OpenAI Whisper (accurate for all accents)
    """
    from twilio.rest import Client
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    # <Pause> waits after the Twilio trial disclaimer + keypress
    # finishOnKey="" — no key ends the recording, only silence/timeout does
    twiml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Pause length="2"/>
    <Say voice="Polly.Joanna">
        Hello {person_name}. This is an automated HR call.
        Please speak about your current availability after the beep.
        Your response will be recorded.
    </Say>
    <Record
        maxLength="60"
        timeout="5"
        finishOnKey=""
        playBeep="true"
    />
    <Say>Thank you for your response. Goodbye.</Say>
</Response>"""

    # Note: removed transcribe="true" — Whisper handles this now

    print(f"\n  📞 Calling {person_name} at {to_number}...")
    print(f"     Tip: Hear trial disclaimer → wait → hear prompt → speak after beep")
    call = client.calls.create(twiml=twiml, to=to_number, from_=TWILIO_FROM_NUMBER)
    print(f"     Call SID: {call.sid}")

    # Poll until call ends
    for _ in range(24):
        time.sleep(5)
        call = client.calls(call.sid).fetch()
        print(f"     Call status: {call.status}")
        if call.status in ("completed", "failed", "busy", "no-answer"):
            break

    if call.status != "completed":
        print(f"     ⚠️  Call ended as: {call.status}")
        return {"call_status": "No Answer", "transcript": ""}

    # Wait for Twilio to finalise the recording
    print("     Waiting for recording to be ready...")
    time.sleep(8)

    # Fetch recording
    recordings = client.recordings.list(call_sid=call.sid, limit=1)
    if not recordings:
        print("     ⚠️  No recording found")
        return {"call_status": "Completed", "transcript": "No speech recorded"}

    rec = recordings[0]
    # Build the direct MP3 download URL
    recording_url = f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_ACCOUNT_SID}/Recordings/{rec.sid}.mp3"
    print(f"     Recording SID: {rec.sid}")

    # Transcribe with Whisper
    transcript = transcribe_with_whisper(recording_url)
    return {"call_status": "Completed", "transcript": transcript}


# ─────────────────────────────────────────────────────────────────
# EMAIL AGENT
# ─────────────────────────────────────────────────────────────────
def send_email(to_email: str, person_name: str) -> dict:
    """Sends personalised email via Gmail SMTP (port 587, STARTTLS)."""
    subject   = f"Hi {person_name} - Employment Status Follow-up"
    body_text = (
        f"Hi {person_name},\n\n"
        "We hope you are doing well.\n\n"
        "We wanted to reach out regarding your current employment status "
        "and gather some information from you.\n\n"
        "Please reply to this email with your availability for a quick chat, "
        "or call us at our HR helpline.\n\n"
        "Best regards,\nHR Team"
    )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = GMAIL_ADDRESS
    msg["To"]      = to_email
    msg.attach(MIMEText(body_text, "plain"))

    print(f"  📧 Sending email to {person_name} ({to_email})...")
    print(f"     Gmail: {GMAIL_ADDRESS} | App password set: {'Yes' if 'xxxx' not in GMAIL_APP_PASSWORD else '❌ NOT SET'}")

    # Try port 587 (STARTTLS) first, fall back to port 465 (SSL) if blocked
    # Error 10060 = connection timed out, usually means port 587 is blocked by network/firewall
    attempts = [
        ("587 STARTTLS", lambda: _send_587(GMAIL_ADDRESS, GMAIL_APP_PASSWORD, to_email, msg)),
        ("465 SSL",      lambda: _send_465(GMAIL_ADDRESS, GMAIL_APP_PASSWORD, to_email, msg)),
    ]

    for label, send_fn in attempts:
        try:
            print(f"     Trying port {label}...")
            send_fn()
            print("     ✅ Email sent successfully")
            return {"email_sent": "Yes", "email_text": body_text}
        except smtplib.SMTPAuthenticationError as e:
            print(f"     ❌ Auth failed: {e}")
            print("     → Use App Password from https://myaccount.google.com/apppasswords")
            return {"email_sent": "No", "email_text": ""}   # no point retrying auth errors
        except Exception as e:
            print(f"     ⚠️  Port {label} failed: {e}")
            continue   # try next port

    print("     ❌ Email failed on all ports")
    return {"email_sent": "No", "email_text": ""}


def _send_587(gmail_address, app_password, to_email, msg):
    """STARTTLS on port 587"""
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(gmail_address, app_password)
        server.sendmail(gmail_address, to_email, msg.as_string())


def _send_465(gmail_address, app_password, to_email, msg):
    """SSL on port 465 — works when port 587 is blocked"""
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as server:
        server.ehlo()
        server.login(gmail_address, app_password)
        server.sendmail(gmail_address, to_email, msg.as_string())


# ─────────────────────────────────────────────────────────────────
# WRITE NEW COLUMNS INTO EXISTING FILE  (preserves all original data + formatting)
# ─────────────────────────────────────────────────────────────────
def save_to_excel(df: pd.DataFrame):
    """
    Opens the original file with openpyxl and ONLY appends new columns.
    - Does NOT touch any existing data, formatting, or column widths.
    - Checks if new columns already exist before adding — safe to call once.
    """
    NEW_COLS = ["STATUS", "EMAIL_SENT", "EMAIL_TEXT", "PHONE_CALL_STATUS", "PHONE_CONVERSATION"]

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Read existing headers from row 1 to avoid duplicate columns
    existing_headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Find where new columns should start — after last existing column
    next_col = ws.max_column + 1
    col_map  = {}   # col_name → column index

    for col_name in NEW_COLS:
        if col_name in existing_headers:
            # Column already exists — reuse it (overwrite values only)
            col_map[col_name] = existing_headers[col_name]
        else:
            # New column — add header at next available position
            ws.cell(row=1, column=next_col, value=col_name)
            col_map[col_name] = next_col
            next_col += 1

    # Write data rows (row 2 onwards) — only write non-empty values
    for col_name in NEW_COLS:
        col_idx = col_map[col_name]
        for row_idx, value in enumerate(df[col_name], start=2):
            if value != "" and value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)
            # Leave cell completely untouched if value is empty

    wb.save(OUTPUT_FILE)
    print(f"  ✅ New columns written to {OUTPUT_FILE} (original data completely untouched)")


# ─────────────────────────────────────────────────────────────────
# MAIN FLOW
# ─────────────────────────────────────────────────────────────────
def process_records():
    print("Reading Excel file...")
    df = pd.read_excel(INPUT_FILE)
    print(f"  Loaded {len(df)} records\n")

    df["STATUS"]             = ""
    df["EMAIL_SENT"]         = ""
    df["EMAIL_TEXT"]         = ""
    df["PHONE_CALL_STATUS"]  = ""
    df["PHONE_CONVERSATION"] = ""  ""

    for idx, row in df.iterrows():
        birth_date = row["BIRTH_DATE"]
        term_date  = row["TERMINATION_DATE"]
        term_code  = row["TERMINATION_CODE"]
        name       = f"{row['FIRST_NAME']} {row['LAST_NAME']}"

        # Diamond 1: Born after 2000-01-01?
        if pd.isna(birth_date) or birth_date <= BIRTH_CUTOFF:
            continue

        is_terminated = (term_code == 152) or (not pd.isna(term_date))

        if not is_terminated:
            # Born >2000 but NOT terminated → Read Next Rec (skip, per flowchart)
            print(f"  ⏭  {name} → Born after 2000 but not terminated, skipping")
            continue

        # Diamond 2: Term Date > 01-01-2023?
        if pd.isna(term_date) or term_date <= TERM_DATE_CUTOFF:
            # ── DQ path: Terminated but BEFORE 2023
            # Flowchart: just capture Name, Phone, Email + Status='DQ' → save to Excel
            # NO email, NO call
            print(f"  ✅ {name} → DisQualified (terminated before 2023)")
            df.at[idx, "STATUS"] = "DisQualified"
            # Per flowchart: DQ path only captures Name/Phone/Email + Status=DQ
            # No email sent, no call made — leave EMAIL_SENT, PHONE_CALL_STATUS blank
        else:
            # ── Call + Email path: Terminated AFTER 2023
            # Flowchart: capture Phone# & Email → AI call → AI email "Hi" → record & save
            print(f"  ✅ {name} → Terminated after 2023, initiating call + email")

            # AI Agent 1: Make the call and collect information
            call_result = make_call_and_transcribe(TEST_PHONE, name)  # swap → row["WORK_PHONE"] for production
            df.at[idx, "PHONE_CALL_STATUS"]  = call_result["call_status"]
            df.at[idx, "PHONE_CONVERSATION"] = call_result["transcript"]

            # AI Agent 2: Email the person "Hi"
            email_result = send_email(TEST_EMAIL, name)   # swap → row["EMAIL"] for production
            df.at[idx, "EMAIL_SENT"] = email_result["email_sent"]
            df.at[idx, "EMAIL_TEXT"] = email_result["email_text"]

    save_to_excel(df)

    print(f"\n{'='*50}")
    print(f"✅ Done! Results written to: {OUTPUT_FILE}")
    print(f"   DisQualified : {(df['STATUS'] == 'DisQualified').sum()}")
    print(f"   Emails sent  : {(df['EMAIL_SENT'] == 'Yes').sum()}")
    print(f"   Calls done   : {(df['PHONE_CALL_STATUS'] == 'Completed').sum()}")
    print(f"   No Answer    : {(df['PHONE_CALL_STATUS'] == 'No Answer').sum()}")
    print(f"{'='*50}")


# ─────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if "--test-email" in sys.argv:
        print("=== EMAIL TEST ===")
        send_email(TEST_EMAIL, "Deepak")
    elif "--test-call" in sys.argv:
        print("=== CALL TEST ===")
        result = make_call_and_transcribe(TEST_PHONE, "Deepak")
        print(f"\nFinal result: {result}")
    else:
        process_records()