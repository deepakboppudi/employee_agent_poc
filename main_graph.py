"""
Employee Agent — LangGraph + LangChain + Groq Whisper
=====================================================
Architecture:
  - LangGraph       : state machine — 4 nodes, each does one thing
  - LangChain @tool : make_call_tool, send_email_tool — exposed as tools for discoverability
  - Groq Whisper    : voice-to-text transcription of recorded calls (free)

Graph nodes:
  filter_node     — rule-based: birth date + termination check
  router_node     — rule-based: term date check (contact vs disqualify)
  contact_node    — calls make_call_tool + send_email_tool directly (no LLM loop)
  disqualify_node — sets STATUS=DisQualified, no call/email
  save_node       — marks record ready for Excel write

Why no LLM in the contact node:
  The contact path is deterministic — always call then email, exactly once.
  Using create_agent here caused an uncontrolled ReAct loop that called
  the tools multiple times per person. Direct invocation is correct here.

Install:
  pip install langchain langgraph langchain-core
              twilio groq pandas openpyxl requests
"""

import os
import sys
import time
import smtplib
import requests
import pandas as pd
from typing import TypedDict
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# LangChain / LangGraph
from langchain_core.tools import tool
from langgraph.graph import StateGraph, END

import dotenv
dotenv.load_dotenv()  # Load credentials from .env file

# All credentials and settings from config.py
from config import (
    TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_FROM_NUMBER,
    GMAIL_ADDRESS, GMAIL_APP_PASSWORD,
    GROQ_API_KEY, BIRTH_CUTOFF, TERM_DATE_CUTOFF,
    INPUT_FILE, OUTPUT_FILE,
)


# ─────────────────────────────────────────────────────────────────
# AGENT STATE — TypedDict flows through every LangGraph node
# ─────────────────────────────────────────────────────────────────
class EmployeeState(TypedDict):
    # Input fields (from Excel row)
    name:       str
    phone:      str
    email:      str
    birth_date: object
    term_date:  object
    term_code:  int
    row_index:  int

    # Output fields (written back to Excel)
    status:             str
    email_sent:         str
    email_text:         str
    phone_call_status:  str
    phone_conversation: str

    # Routing — tells conditional edges which node to go to
    next: str


# ─────────────────────────────────────────────────────────────────
# TOOL RESULTS STORE
# LangChain tools are stateless — this lets them write results
# back so llm_agent_node can pick them up after tool execution
# ─────────────────────────────────────────────────────────────────
_tool_results: dict = {}


# ─────────────────────────────────────────────────────────────────
# LANGCHAIN TOOLS
# @tool exposes these to the LLM — it reads the docstring to
# understand what each tool does and when to call it
# ─────────────────────────────────────────────────────────────────
@tool
def make_call_tool(phone: str, name: str) -> str:
    """
    Place an outbound phone call to the employee, record their spoken
    response, and transcribe it to text using Groq Whisper.
    Use this when the employee was terminated AFTER 2023-01-01.
    Returns a summary of the call outcome and transcript.
    """
    from twilio.rest import Client
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    twiml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="Polly.Joanna">
        Hello {name}. This is an automated HR call.
        Please speak about your current availability after the beep.
        Your response will be recorded.
    </Say>
    <Record maxLength="60" timeout="5" finishOnKey="" playBeep="true"/>
</Response>"""

    print(f"     [make_call_tool] Calling {name} at {phone}...")

    # timeout=30 tells Twilio to stop ringing after 30 seconds if not answered
    # This prevents the call from blocking the loop for 60+ seconds
    call = client.calls.create(
        twiml=twiml,
        to=phone,
        from_=TWILIO_FROM_NUMBER,
        timeout=30,
    )

    # Poll for max 45 seconds (9 x 5s) — enough time for 30s ring + connect + short recording
    for _ in range(9):
        time.sleep(5)
        call = client.calls(call.sid).fetch()
        print(f"     Call status: {call.status}")
        if call.status in ("completed", "failed", "busy", "no-answer"):
            break

    # If still ringing/in-progress after our poll window, treat as no answer
    if call.status in ("queued", "ringing", "in-progress"):
        print(f"     Call timed out in poll window — treating as No Answer")
        _tool_results["phone_call_status"]  = "No Answer"
        _tool_results["phone_conversation"] = ""
        return "Call timed out — no answer"

    status_map = {
        "completed": "Completed",
        "no-answer": "No Answer",
        "busy":      "Busy",
        "failed":    "Failed",
    }
    call_status = status_map.get(call.status, call.status.title())

    if call.status != "completed":
        _tool_results["phone_call_status"]  = call_status
        _tool_results["phone_conversation"] = ""
        return f"Call ended with status: {call_status}"

    time.sleep(8)
    recordings = client.recordings.list(call_sid=call.sid, limit=1)
    if not recordings:
        _tool_results["phone_call_status"]  = "Completed"
        _tool_results["phone_conversation"] = "No speech recorded"
        return "Call completed but no recording found"

    rec      = recordings[0]
    duration = int(rec.duration) if rec.duration else 0
    if duration < 2:
        _tool_results["phone_call_status"]  = "Completed"
        _tool_results["phone_conversation"] = "No response spoken"
        return "Call completed but no response spoken"

    recording_url = (
        f"https://api.twilio.com/2010-04-01/Accounts/"
        f"{TWILIO_ACCOUNT_SID}/Recordings/{rec.sid}.mp3"
    )
    transcript = _transcribe(recording_url)
    _tool_results["phone_call_status"]  = "Completed"
    _tool_results["phone_conversation"] = transcript
    print(f"     ✅ Phone conversation captured: {transcript}")
    return f"Call completed. Transcript: {transcript}"


@tool
def send_email_tool(to_email: str, name: str) -> str:
    """
    Send a personalised HR follow-up email to the employee via Gmail SMTP.
    Auto-falls back from port 587 (STARTTLS) to port 465 (SSL) if blocked.
    Use this when the employee was terminated AFTER 2023-01-01.
    Returns confirmation of whether the email was sent successfully.
    """
    subject   = f"Hi {name} - Employment Status Follow-up"
    body_text = (
        f"Hi {name},\n\n"
        "We hope you are doing well.\n\n"
        "We wanted to reach out regarding your current employment status "
        "and gather some information from you.\n\n"
        "Please reply to this email with your availability for a quick chat, "
        "or call us at our HR helpline.\n\n"
        "Best regards,\nHR Team"
    )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = GMAIL_ADDRESS
    msg["To"]      = to_email
    msg.attach(MIMEText(body_text, "plain"))

    print(f"     [send_email_tool] Emailing {name} at {to_email}...")

    for port, use_ssl in [(587, False), (465, True)]:
        try:
            if use_ssl:
                server = smtplib.SMTP_SSL("smtp.gmail.com", port, timeout=15)
            else:
                server = smtplib.SMTP("smtp.gmail.com", port, timeout=15)
                server.ehlo()
                server.starttls()
            server.ehlo()
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_ADDRESS, to_email, msg.as_string())
            server.quit()
            print(f"     Email sent via port {port}")
            _tool_results["email_sent"] = "Yes"
            _tool_results["email_text"] = body_text
            return f"Email sent successfully to {to_email}"
        except smtplib.SMTPAuthenticationError as e:
            _tool_results["email_sent"] = "No"
            return f"Email auth failed: {e}"
        except Exception:
            continue

    _tool_results["email_sent"] = "No"
    return "Email failed on all ports"


@tool
def mark_disqualified_tool(name: str) -> str:
    """
    Mark the employee as DisQualified in the HR system.
    Use this when the employee was terminated BEFORE or ON 2023-01-01.
    No call or email is sent — only the status is recorded.
    Returns confirmation that the employee has been marked DisQualified.
    """
    print(f"     [mark_disqualified_tool] Marking {name} as DisQualified")
    _tool_results["status"] = "DisQualified"
    return f"{name} marked as DisQualified"


# ─────────────────────────────────────────────────────────────────
# GROQ WHISPER — voice-to-text transcription
# ─────────────────────────────────────────────────────────────────
def _transcribe(recording_url: str) -> str:
    from groq import Groq
    client = Groq(api_key=GROQ_API_KEY)

    resp = requests.get(
        recording_url,
        auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
        timeout=30
    )
    if resp.status_code != 200:
        return "Recording download failed"

    tmp = "tmp_recording.mp3"
    with open(tmp, "wb") as f:
        f.write(resp.content)

    try:
        for attempt in range(1, 4):
            try:
                with open(tmp, "rb") as audio:
                    result = client.audio.transcriptions.create(
                        model="whisper-large-v3",
                        file=audio,
                        language="en"
                    )
                return result.text
            except Exception as e:
                if "503" in str(e) and attempt < 3:
                    print(f"     Groq 503, retrying in {attempt * 10}s...")
                    time.sleep(attempt * 10)
                else:
                    return f"Transcription error: {e}"
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)

    return "Transcription failed"


# ─────────────────────────────────────────────────────────────────
# LANGGRAPH NODES
# ─────────────────────────────────────────────────────────────────
def filter_node(state: EmployeeState) -> EmployeeState:
    """
    Node 1 — Birth date + termination check (rule-based, no LLM).
    Skips ineligible records before any LLM cost is incurred.
    """
    if pd.isna(state["birth_date"]) or state["birth_date"] <= BIRTH_CUTOFF:
        # print(f"  -- {state['name']} | born before 2000, skip")
        return {**state, "next": "skip"}

    is_terminated = (state["term_code"] == 152) or (state["term_date"] is not None)
    if not is_terminated:
        # print(f"  -- {state['name']} | not terminated, skip")
        return {**state, "next": "skip"}

    return {**state, "next": "router"}


def router_node(state: EmployeeState) -> EmployeeState:
    """
    Node 2 — Term date check (rule-based, no LLM).
    Separating this from filter_node ensures each node does one thing only,
    and the LLM is only invoked for the correct path.
    """
    if state["term_date"] and state["term_date"] > TERM_DATE_CUTOFF:
        print(f"  >> {state['name']} | terminated after 2023 → calling + emailing")
        return {**state, "next": "llm_agent"}

    print(f"  DQ {state['name']} | terminated before 2023 → disqualify")
    return {**state, "next": "disqualify"}


def disqualify_node(state: EmployeeState) -> EmployeeState:
    """
    Node 3a — DQ path: set Status=DisQualified only. No call, no email.
    Bypasses the LLM entirely — no point spending tokens on a deterministic outcome.
    """
    return {
        **state,
        "status":            "DisQualified",
        "phone_call_status": "Not Applicable",
        "next":              "save",
    }


def contact_node(state: EmployeeState) -> EmployeeState:
    """
    Node 3b — Contact path (term date > 2023).
    Calls make_call_tool then send_email_tool directly — exactly once each.
    No LLM loop here: the routing decision was already made by router_node.
    Direct invocation prevents the ReAct agent from calling tools repeatedly.
    """
    global _tool_results
    # Reset ALL keys explicitly — prevents stale data from previous employee
    # leaking into this record if a tool path is skipped (e.g. No Answer)
    _tool_results = {
        "phone_call_status":  "",
        "phone_conversation": "",
        "email_sent":         "",
        "email_text":         "",
        "status":             "",
    }

    print(f"\n  >> Contacting {state['name']}")

    # Step 1: Phone call (tool invoked directly — not via LLM)
    call_result = make_call_tool.invoke({
        "phone": state["phone"],
        "name":  state["name"],
    })
    print(f"     Phone call status : {_tool_results.get('phone_call_status', 'unknown')}")
    print(f"     Transcript        : {_tool_results.get('phone_conversation', '(none)')}")

    # Step 2: Email (tool invoked directly — not via LLM)
    email_result = send_email_tool.invoke({
        "to_email": state["email"],
        "name":     state["name"],
    })
    print(f"     Email sent        : {_tool_results.get('email_sent', 'No')}")

    return {
        **state,
        "status":             "Contacted",
        "email_sent":         _tool_results.get("email_sent", "No"),
        "email_text":         _tool_results.get("email_text", ""),
        "phone_call_status":  _tool_results.get("phone_call_status", ""),
        "phone_conversation": _tool_results.get("phone_conversation", ""),
        "next": "save",
    }


def save_node(state: EmployeeState) -> EmployeeState:
    return {**state, "next": END}


# ─────────────────────────────────────────────────────────────────
# LANGGRAPH GRAPH
# ─────────────────────────────────────────────────────────────────
def build_graph():
    """
    Four-node graph — each node does exactly one thing:
      filter_node     — birth date + termination check (skip or pass)
      router_node     — term date check (contact or disqualify)
      disqualify_node — set STATUS=DisQualified, no LLM needed
      contact_node    — calls make_call_tool + send_email_tool directly (once each)
      save_node       — marks state ready for Excel write
    """
    graph = StateGraph(EmployeeState)

    graph.add_node("filter",     filter_node)
    graph.add_node("router",     router_node)
    graph.add_node("disqualify", disqualify_node)
    graph.add_node("contact",    contact_node)
    graph.add_node("save",       save_node)

    graph.set_entry_point("filter")

    graph.add_conditional_edges(
        "filter",
        lambda s: s["next"],
        {"skip": END, "router": "router"}
    )
    graph.add_conditional_edges(
        "router",
        lambda s: s["next"],
        {"llm_agent": "contact",   "disqualify": "disqualify"}
    )
    graph.add_edge("disqualify", "save")
    graph.add_edge("contact",    "save")
    graph.add_edge("save",       END)

    return graph.compile()


# ─────────────────────────────────────────────────────────────────
# EXCEL — append-only, never touches original columns
# ─────────────────────────────────────────────────────────────────
def save_results_to_excel(results: list):
    NEW_COLS = ["STATUS", "EMAIL_SENT", "EMAIL_TEXT",
                "PHONE_CALL_STATUS", "PHONE_CONVERSATION"]

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    existing = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }

    next_col = ws.max_column + 1
    col_map  = {}
    for col_name in NEW_COLS:
        if col_name in existing:
            col_map[col_name] = existing[col_name]
        else:
            ws.cell(row=1, column=next_col, value=col_name)
            col_map[col_name] = next_col
            next_col += 1

    for res in results:
        # Only write rows that were actually processed (contacted or DQ)
        # Skipped rows (born before 2000 / not terminated) have all fields empty
        was_processed = res.get("status") != "" or res.get("phone_call_status") != ""
        if not was_processed:
            continue

        excel_row = res["row_index"] + 2
        for col_name, key in [
            ("STATUS",             "status"),
            ("EMAIL_SENT",         "email_sent"),
            ("EMAIL_TEXT",         "email_text"),
            ("PHONE_CALL_STATUS",  "phone_call_status"),
            ("PHONE_CONVERSATION", "phone_conversation"),
        ]:
            value = res.get(key, "")
            # Write value always — None clears old stale cell content (e.g. No Answer clears old transcript)
            ws.cell(row=excel_row, column=col_map[col_name], value=value if value else None)

    wb.save(OUTPUT_FILE)
    print(f"\n  Results written to {OUTPUT_FILE}")


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────
def process_records():
    print("Reading Excel file...")
    df = pd.read_excel(INPUT_FILE)
    print(f"  Loaded {len(df)} records\n")

    agent   = build_graph()
    results = []

    for idx, row in df.iterrows():
        term_date = row["TERMINATION_DATE"]
        state: EmployeeState = {
            "name":       f"{row['FIRST_NAME']} {row['LAST_NAME']}",
            "phone":      str(row["WORK_PHONE"]).strip(),
            "email":      str(row["EMAIL"]).strip(),
            "birth_date": row["BIRTH_DATE"],
            "term_date":  None if pd.isna(term_date) else term_date,
            "term_code":  int(row["TERMINATION_CODE"]),
            "row_index":  int(idx),
            "status":             "",
            "email_sent":         "",
            "email_text":         "",
            "phone_call_status":  "",
            "phone_conversation": "",
            "next":               "",
        }
        final = agent.invoke(state)
        results.append(final)

    save_results_to_excel(results)

    dq    = sum(1 for r in results if r.get("status") == "DisQualified")
    sent  = sum(1 for r in results if r.get("email_sent") == "Yes")
    calls = sum(1 for r in results if r.get("phone_call_status") == "Completed")
    na    = sum(1 for r in results if r.get("phone_call_status") == "No Answer")

    print(f"\n{'='*50}")
    print(f"Done! Output: {OUTPUT_FILE}")
    print(f"  DisQualified : {dq}")
    print(f"  Emails sent  : {sent}")
    print(f"  Calls done   : {calls}")
    print(f"  No Answer    : {na}")
    print(f"{'='*50}")


# ─────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if "--test-email" in sys.argv:
        result = send_email_tool.invoke({"to_email": "test@example.com", "name": "Test User"})
        print(result)
    elif "--test-call" in sys.argv:
        phone = input("Enter phone number (+91XXXXXXXXXX): ").strip()
        result = make_call_tool.invoke({"phone": phone, "name": "Test User"})
        print(result)
    else:
        process_records()